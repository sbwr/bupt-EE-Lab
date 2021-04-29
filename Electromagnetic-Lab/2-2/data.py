from openpyxl import load_workbook
import sys

path = sys.path[0]+'\\data.xlsx'
workbook = load_workbook(path)              # 找到需要xlsx文件的位置
booksheet = workbook.active                 # 获取当前活跃的sheet,默认是第一个sheet

# 获取

#获取sheet页的行数据
rows = booksheet.rows

#获取sheet页的列数据
columns = booksheet.columns

i = 0
# 迭代所有的行
for row in rows:
    i = i + 1
    line = [col.value for col in row]
    cell_data_1 = booksheet.cell(row=i, column=3).value               #获取第i行1 列的数据
    cell_data_2 = booksheet.cell(row=i, column=4).value               #获取第i行 2 列的数据
    cell_data_3 = booksheet.cell(row=i, column=8).value                   #获取第i行 3 列的数据
    cell_data_4 = booksheet.cell(row=i, column=18).value                   #获取第i行 4 列的数据
    print (cell_data_1, cell_data_2, cell_data_3, cell_data_4)